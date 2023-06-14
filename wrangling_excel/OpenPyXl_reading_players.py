# pip install openpyxl
from openpyxl import Workbook

# import more modules
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Read in the excel sheet
wb = load_workbook(filename = 'OpenPyXl_creating_players.xlsx')
ws = wb['Players info']

# Create a new column:
ws['G1'] = 'BMI'

# for each row
for r in range(2, ws.max_row + 1):
    # take the value in column 6
    weight = ws.cell(row = r, column = 6).value
    # take the value in column 4
    height = ws.cell(row = r, column = 4).value

    # apply a function to calculate bmi
    bmi = round(weight / (height ** 2), 2)

    # create a new column on column 7
    ws.cell(row = r, column = 7).value = bmi

wb.save('OpenPyXl_creating_players.xlsx')
