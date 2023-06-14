

# pip install openpyxl

from openpyxl import Workbook
wb = Workbook()

# create an empty excel file
wb.save('OpenPyXl_creating_players.xlsx')


# import more modules
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def str_to_int_or_float(value):
    if isinstance(value, bool):
        return value
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value

# Load workbook, select active sheet and rename it:
wb = Workbook()
ws = wb.active
ws.title = 'Players info'


# Add the headings separately
ws.append(['First name', 'Last name', 'Jersey', 'Height [mts]', 'NBA debut year', 'Weight [kgs]'])

# use nba_players as datasource:
nba_players = [{'firstName': 'LeBron', 'lastName': 'James', 'jersey': '2', 'heightMeters': '2.03', 'nbaDebutYear': '2003', 'weightKilograms': '113.4'}, {'firstName': 'LaMarcus', 'lastName': 'Aldridge', 'jersey': '12', 'heightMeters': '2.11', 'nbaDebutYear': '2006', 'weightKilograms': '117.9'}, {'firstName': 'Kawhi', 'lastName': 'Leonard', 'jersey': '2', 'heightMeters': '2.01', 'nbaDebutYear': '2011', 'weightKilograms': '104.3'}, {'firstName': 'Jabari', 'lastName': 'Parker', 'jersey': '2', 'heightMeters': '2.03', 'nbaDebutYear': '2014', 'weightKilograms': '111.1'}]

for player in nba_players:
    ws.append(list(map(str_to_int_or_float, player.values())))


# Create an Excel table that starts at A1 and ends on the last non-empty cell:
last_cell = ws.cell(row = ws.max_row, column = ws.max_column).coordinate
player_table = Table(displayName = 'PlayerTable', ref = 'A1:{}'.format(last_cell))


# Style the table and add it to the spreadsheet. Note that you can use any of the table styles available in your Excel version (we will use Table Style Medium 6 here) without spaces in its name. Finally, save changes.
style = TableStyleInfo(name = 'TableStyleMedium6', showRowStripes=True)
player_table.tableStyleInfo = style
ws.add_table(player_table)
wb.save('OpenPyXl_creating_players.xlsx')